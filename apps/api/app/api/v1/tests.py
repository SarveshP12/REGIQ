"""Test Repository CRUD endpoints — /api/v1/tests."""

import io
import math
import uuid
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, func, or_, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit import log_audit
from app.core.database import get_db
from app.core.mongodb import get_mongo_db
from app.core.security import get_current_user, require_permissions
from app.models.test_case import TestCase
from app.models.user import User
from app.schemas.test_case import (
    ImportResult,
    TestCaseCreate,
    TestCaseListResponse,
    TestCaseResponse,
    TestCaseUpdate,
    TestCaseVersionResponse,
    DuplicateCheckRequest,
)

router = APIRouter()


# ── List with pagination, filtering, search ─────────────


@router.get("/", response_model=TestCaseListResponse, dependencies=[Depends(require_permissions("tests:read"))])
async def list_tests(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status_filter: Optional[str] = Query(None, alias="status"),
    criticality: Optional[str] = None,
    automation_flag: Optional[str] = None,
    business_process_id: Optional[uuid.UUID] = None,
    search: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List test cases with pagination, filtering, and full-text search."""
    query = select(TestCase).where(
        TestCase.tenant_id == current_user.tenant_id,
        TestCase.archived_at.is_(None),
    )

    # Filters
    if status_filter:
        query = query.where(TestCase.status == status_filter)
    if criticality:
        query = query.where(TestCase.criticality == criticality)
    if automation_flag:
        query = query.where(TestCase.automation_flag == automation_flag)
    if business_process_id:
        query = query.where(TestCase.business_process_id == business_process_id)

    # Full-text search using PostgreSQL tsvector
    if search:
        search_condition = or_(
            TestCase.title.ilike(f"%{search}%"),
            TestCase.description.ilike(f"%{search}%"),
        )
        query = query.where(search_condition)

    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total_result = await db.execute(count_query)
    total = total_result.scalar() or 0

    # Paginate
    offset = (page - 1) * page_size
    query = query.order_by(TestCase.updated_at.desc()).offset(offset).limit(page_size)

    result = await db.execute(query)
    items = result.scalars().all()

    return TestCaseListResponse(
        items=[TestCaseResponse.model_validate(tc) for tc in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=math.ceil(total / page_size) if total > 0 else 0,
    )


# ── Create ────────────────────────────────────────────────


@router.post("/", response_model=TestCaseResponse, status_code=201, dependencies=[Depends(require_permissions("tests:write"))])
async def create_test(
    body: TestCaseCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new test case."""
    tc = TestCase(
        **body.model_dump(),
        tenant_id=current_user.tenant_id,
        created_by=current_user.id,
    )
    
    # Run auto-classification
    try:
        from app.core.ai.classifier import classify_test_case
        steps_text = _steps_to_text(tc.steps)
        cls_res = classify_test_case(
            title=tc.title or "",
            description=tc.description or "",
            steps=steps_text,
        )
        tc.ai_business_process = cls_res.business_process
        tc.ai_test_case_type = cls_res.test_case_type
        tc.ai_dependency_class = cls_res.dependency_class
        tc.ai_automation_feasibility = cls_res.automation_feasibility
        tc.ai_execution_frequency = cls_res.execution_frequency
        tc.ai_confidence_scores = cls_res.confidence_scores
        tc.ai_needs_review = cls_res.needs_review
        tc.ai_model_version = cls_res.model_version
        tc.ai_classified_at = datetime.now(timezone.utc)
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Auto-classification failed for new test case: %s", e)

    # Run criticality scoring
    try:
        from app.core.ai.criticality import calculate_criticality
        steps_text = _steps_to_text(tc.steps)
        crit_res = calculate_criticality(
            title=tc.title or "",
            description=tc.description or "",
            steps=steps_text,
        )
        tc.criticality_score = crit_res["score"]
        tc.ai_criticality_level = crit_res["category"]
        tc.criticality = crit_res["category"]
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Criticality scoring failed for new test case: %s", e)

    # Generate pgvector embedding
    try:
        from app.core.ai.embedding import embedding_service
        tc.embedding = embedding_service.generate_embedding(tc.title + " " + (tc.description or ""))
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Embedding generation failed for new test case: %s", e)

    db.add(tc)
    await db.flush()
    await db.refresh(tc)

    # Link test case in Neo4j
    try:
        from app.services.graph_builder import GraphBuilderService
        steps_text = _steps_to_text(tc.steps)
        builder = GraphBuilderService()
        builder.link_test_case(
            tc_id=str(tc.id),
            title=tc.title or "",
            description=tc.description or "",
            steps_text=steps_text,
            tags=tc.type_tags or []
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Neo4j linking failed for new test case: %s", e)

    # Save initial version snapshot to MongoDB
    mongo = get_mongo_db()
    await mongo.test_case_versions.insert_one({
        "test_case_id": str(tc.id),
        "version": 1,
        "snapshot": body.model_dump(mode="json"),
        "changed_fields": [],
        "changed_by": str(current_user.id),
        "timestamp": datetime.now(timezone.utc),
    })

    # Audit log
    await log_audit(
        db, current_user.id, "create", "test_case",
        resource_id=tc.id, tenant_id=current_user.tenant_id,
        details={"title": tc.title},
    )

    return TestCaseResponse.model_validate(tc)


# ── Get detail ────────────────────────────────────────────


@router.get("/{test_id}", response_model=TestCaseResponse, dependencies=[Depends(require_permissions("tests:read"))])
async def get_test(
    test_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get full test case details."""
    tc = await _get_test_or_404(db, test_id, current_user.tenant_id)
    return TestCaseResponse.model_validate(tc)


# ── Update (creates new version) ──────────────────────────


@router.put("/{test_id}", response_model=TestCaseResponse, dependencies=[Depends(require_permissions("tests:write"))])
async def update_test(
    test_id: uuid.UUID,
    body: TestCaseUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update a test case — creates a new version snapshot in MongoDB."""
    tc = await _get_test_or_404(db, test_id, current_user.tenant_id)

    update_data = body.model_dump(exclude_unset=True)
    changed_fields = list(update_data.keys())

    for field, value in update_data.items():
        setattr(tc, field, value)

    # Run AI services if key text fields are updated
    if any(f in changed_fields for f in ["title", "description", "steps", "preconditions", "expected_results"]):
        steps_text = _steps_to_text(tc.steps)
        
        # Run classification
        try:
            from app.core.ai.classifier import classify_test_case
            cls_res = classify_test_case(
                title=tc.title or "",
                description=tc.description or "",
                steps=steps_text,
            )
            tc.ai_business_process = cls_res.business_process
            tc.ai_test_case_type = cls_res.test_case_type
            tc.ai_dependency_class = cls_res.dependency_class
            tc.ai_automation_feasibility = cls_res.automation_feasibility
            tc.ai_execution_frequency = cls_res.execution_frequency
            tc.ai_confidence_scores = cls_res.confidence_scores
            tc.ai_needs_review = cls_res.needs_review
            tc.ai_model_version = cls_res.model_version
            tc.ai_classified_at = datetime.now(timezone.utc)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Auto-classification failed for updated test case: %s", e)

        # Run criticality
        try:
            from app.core.ai.criticality import calculate_criticality
            crit_res = calculate_criticality(
                title=tc.title or "",
                description=tc.description or "",
                steps=steps_text,
            )
            tc.criticality_score = crit_res["score"]
            tc.ai_criticality_level = crit_res["category"]
            tc.criticality = crit_res["category"]
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Criticality scoring failed for updated test case: %s", e)

        # Generate embedding
        try:
            from app.core.ai.embedding import embedding_service
            tc.embedding = embedding_service.generate_embedding(tc.title + " " + (tc.description or ""))
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Embedding generation failed for updated test case: %s", e)

        # Link/Update test case in Neo4j
        try:
            from app.services.graph_builder import GraphBuilderService
            builder = GraphBuilderService()
            builder.link_test_case(
                tc_id=str(tc.id),
                title=tc.title or "",
                description=tc.description or "",
                steps_text=steps_text,
                tags=tc.type_tags or []
            )
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Neo4j linking failed for updated test case: %s", e)

    tc.version += 1
    tc.updated_at = datetime.now(timezone.utc)
    await db.flush()
    await db.refresh(tc)

    # Save version snapshot to MongoDB
    mongo = get_mongo_db()
    snapshot = TestCaseResponse.model_validate(tc).model_dump(mode="json")
    await mongo.test_case_versions.insert_one({
        "test_case_id": str(tc.id),
        "version": tc.version,
        "snapshot": snapshot,
        "changed_fields": changed_fields,
        "changed_by": str(current_user.id),
        "timestamp": datetime.now(timezone.utc),
    })

    await log_audit(
        db, current_user.id, "update", "test_case",
        resource_id=tc.id, tenant_id=current_user.tenant_id,
        details={"changed_fields": changed_fields, "version": tc.version},
    )

    return TestCaseResponse.model_validate(tc)


# ── Soft delete (archive) ────────────────────────────────


@router.delete("/{test_id}", status_code=204, dependencies=[Depends(require_permissions("tests:delete"))])
async def delete_test(
    test_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Soft-delete (archive) a test case."""
    tc = await _get_test_or_404(db, test_id, current_user.tenant_id)
    tc.archived_at = datetime.now(timezone.utc)
    tc.status = "archived"
    await db.flush()

    await log_audit(
        db, current_user.id, "archive", "test_case",
        resource_id=tc.id, tenant_id=current_user.tenant_id,
    )


# ── Clone ─────────────────────────────────────────────────


@router.post("/{test_id}/clone", response_model=TestCaseResponse, status_code=201, dependencies=[Depends(require_permissions("tests:write"))])
async def clone_test(
    test_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Clone a test case with a new ID and version 1."""
    source = await _get_test_or_404(db, test_id, current_user.tenant_id)

    clone = TestCase(
        title=f"{source.title} (Clone)",
        description=source.description,
        steps=source.steps,
        preconditions=source.preconditions,
        expected_results=source.expected_results,
        format_type=source.format_type,
        status="draft",
        criticality=source.criticality,
        type_tags=source.type_tags,
        automation_flag=source.automation_flag,
        business_process_id=source.business_process_id,
        tenant_id=current_user.tenant_id,
        created_by=current_user.id,
        version=1,
    )
    db.add(clone)
    await db.flush()
    await db.refresh(clone)

    await log_audit(
        db, current_user.id, "clone", "test_case",
        resource_id=clone.id, tenant_id=current_user.tenant_id,
        details={"source_id": str(test_id)},
    )

    return TestCaseResponse.model_validate(clone)


# ── Version history ───────────────────────────────────────


@router.get("/{test_id}/versions", dependencies=[Depends(require_permissions("tests:read"))])
async def get_versions(
    test_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get version history for a test case from MongoDB."""
    # Verify access
    await _get_test_or_404(db, test_id, current_user.tenant_id)

    mongo = get_mongo_db()
    cursor = mongo.test_case_versions.find(
        {"test_case_id": str(test_id)},
        {"_id": 0},
    ).sort("version", -1)

    versions = await cursor.to_list(length=100)
    return versions


# ── Bulk import from Excel/CSV ────────────────────────────


@router.post("/import", response_model=ImportResult, dependencies=[Depends(require_permissions("tests:import"))])
async def import_tests(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Bulk import test cases from Excel (.xlsx) or CSV."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in ("xlsx", "csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported")

    content = await file.read()
    rows: list[dict] = []

    if ext == "xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Empty workbook")

        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{j}" for j, h in enumerate(row)]
                continue
            row_dict = {}
            for j, val in enumerate(row):
                if j < len(headers) and val is not None:
                    row_dict[headers[j]] = str(val)
            if row_dict:
                rows.append(row_dict)
    else:
        import csv as csv_mod

        text_content = content.decode("utf-8-sig")
        reader = csv_mod.DictReader(io.StringIO(text_content))
        for row in reader:
            rows.append({k.strip().lower().replace(" ", "_"): v.strip() for k, v in row.items() if v and v.strip()})

    # Import rows into DB
    imported = 0
    errors: list[dict] = []
    mongo = get_mongo_db()

    for idx, row in enumerate(rows, start=2):
        title = row.get("title") or row.get("test_case_name") or row.get("name")
        if not title:
            errors.append({"row": idx, "error": "Missing title"})
            continue

        tc = TestCase(
            title=title[:500],
            description=row.get("description", ""),
            preconditions=row.get("preconditions"),
            expected_results=row.get("expected_results") or row.get("expected_result"),
            format_type=row.get("format_type", "structured"),
            status=row.get("status", "draft"),
            criticality=row.get("criticality", "medium"),
            type_tags=[t.strip() for t in row.get("type_tags", "").split(",") if t.strip()] or [],
            automation_flag=row.get("automation_flag", "manual"),
            tenant_id=current_user.tenant_id,
            created_by=current_user.id,
        )

        # Parse steps if present
        steps_raw = row.get("steps")
        if steps_raw:
            try:
                import json
                tc.steps = json.loads(steps_raw)
            except (json.JSONDecodeError, TypeError):
                tc.steps = [{"step": 1, "action": steps_raw}]

        # Run auto-classification
        try:
            from app.core.ai.classifier import classify_test_case
            steps_text = _steps_to_text(tc.steps)
            cls_res = classify_test_case(
                title=tc.title or "",
                description=tc.description or "",
                steps=steps_text,
            )
            tc.ai_business_process = cls_res.business_process
            tc.ai_test_case_type = cls_res.test_case_type
            tc.ai_dependency_class = cls_res.dependency_class
            tc.ai_automation_feasibility = cls_res.automation_feasibility
            tc.ai_execution_frequency = cls_res.execution_frequency
            tc.ai_confidence_scores = cls_res.confidence_scores
            tc.ai_needs_review = cls_res.needs_review
            tc.ai_model_version = cls_res.model_version
            tc.ai_classified_at = datetime.now(timezone.utc)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Auto-classification failed on import: %s", e)

        # Run criticality scoring
        try:
            from app.core.ai.criticality import calculate_criticality
            steps_text = _steps_to_text(tc.steps)
            crit_res = calculate_criticality(
                title=tc.title or "",
                description=tc.description or "",
                steps=steps_text,
            )
            tc.criticality_score = crit_res["score"]
            tc.ai_criticality_level = crit_res["category"]
            tc.criticality = crit_res["category"]
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Criticality scoring failed on import: %s", e)

        # Generate pgvector embedding
        try:
            from app.core.ai.embedding import embedding_service
            tc.embedding = embedding_service.generate_embedding(tc.title + " " + (tc.description or ""))
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Embedding generation failed on import: %s", e)

        db.add(tc)
        await db.flush()

        # Link imported test case in Neo4j
        try:
            from app.services.graph_builder import GraphBuilderService
            steps_text = _steps_to_text(tc.steps)
            builder = GraphBuilderService()
            builder.link_test_case(
                tc_id=str(tc.id),
                title=tc.title or "",
                description=tc.description or "",
                steps_text=steps_text,
                tags=tc.type_tags or []
            )
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning("Neo4j linking failed on import: %s", e)

        # Save initial version to MongoDB
        await mongo.test_case_versions.insert_one({
            "test_case_id": str(tc.id),
            "version": 1,
            "snapshot": {"title": tc.title, "description": tc.description, "status": tc.status},
            "changed_fields": [],
            "changed_by": str(current_user.id),
            "timestamp": datetime.now(timezone.utc),
        })

        imported += 1

    await log_audit(
        db, current_user.id, "import", "test_case",
        tenant_id=current_user.tenant_id,
        details={"total_rows": len(rows), "imported": imported, "errors_count": len(errors)},
    )

    return ImportResult(
        total_rows=len(rows),
        imported=imported,
        skipped=len(errors),
        errors=errors,
    )


# ── Export to Excel ───────────────────────────────────────


@router.get("/export", dependencies=[Depends(require_permissions("tests:export"))])
async def export_tests(
    format: str = Query("xlsx", pattern=r"^(xlsx|csv)$"),
    status_filter: Optional[str] = Query(None, alias="status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export test cases to Excel or CSV."""
    query = select(TestCase).where(
        TestCase.tenant_id == current_user.tenant_id,
        TestCase.archived_at.is_(None),
    )
    if status_filter:
        query = query.where(TestCase.status == status_filter)

    result = await db.execute(query.order_by(TestCase.created_at.desc()))
    test_cases = result.scalars().all()

    if format == "csv":
        import csv as csv_mod

        output = io.StringIO()
        writer = csv_mod.writer(output)
        writer.writerow(["Title", "Description", "Status", "Criticality", "Automation", "Format", "Version", "Created At"])
        for tc in test_cases:
            writer.writerow([tc.title, tc.description or "", tc.status, tc.criticality, tc.automation_flag, tc.format_type, tc.version, tc.created_at.isoformat()])
        content = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=test_cases.csv"},
        )
    else:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        ws.append(["Title", "Description", "Status", "Criticality", "Automation", "Format", "Version", "Created At"])
        for tc in test_cases:
            ws.append([tc.title, tc.description or "", tc.status, tc.criticality, tc.automation_flag, tc.format_type, tc.version, tc.created_at.isoformat()])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=test_cases.xlsx"},
        )


# ── Helpers ───────────────────────────────────────────────


async def _get_test_or_404(db: AsyncSession, test_id: uuid.UUID, tenant_id: uuid.UUID) -> TestCase:
    result = await db.execute(
        select(TestCase).where(TestCase.id == test_id, TestCase.tenant_id == tenant_id)
    )
    tc = result.scalar_one_or_none()
    if tc is None:
        raise HTTPException(status_code=404, detail="Test case not found")
    return tc


def _steps_to_text(steps) -> str:
    if not steps:
        return ""
    if isinstance(steps, list):
        return " ".join(str(s) for s in steps)
    return str(steps)


# ── Similarity and Duplicate Detection ──────────────────────


@router.get("/similar/{test_id}", response_model=list[TestCaseResponse], dependencies=[Depends(require_permissions("tests:read"))])
@router.get("/{test_id}/similar", response_model=list[TestCaseResponse], dependencies=[Depends(require_permissions("tests:read"))])
async def get_similar_tests(
    test_id: uuid.UUID,
    limit: int = Query(5, ge=1, le=20),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Find similar test cases using cosine similarity on pgvector embeddings."""
    tc = await _get_test_or_404(db, test_id, current_user.tenant_id)
    if not tc.embedding:
        return []

    try:
        # Cosine distance operator is <=> in SQL
        distance_expr = TestCase.embedding.cosine_distance(tc.embedding)
        stmt = (
            select(TestCase)
            .where(
                TestCase.tenant_id == current_user.tenant_id,
                TestCase.id != test_id,
                TestCase.embedding.isnot(None),
                TestCase.archived_at.is_(None)
            )
            .order_by(distance_expr)
            .limit(limit)
        )
        result = await db.execute(stmt)
        similar_cases = result.scalars().all()
        return [TestCaseResponse.model_validate(sc) for sc in similar_cases]
    except Exception as e:
        # Fallback in case of database or pgvector issues
        import logging
        logging.getLogger(__name__).warning("Similarity query failed, fallback to title search: %s", e)
        stmt = (
            select(TestCase)
            .where(
                TestCase.tenant_id == current_user.tenant_id,
                TestCase.id != test_id,
                TestCase.title.ilike(f"%{tc.title}%"),
                TestCase.archived_at.is_(None)
            )
            .limit(limit)
        )
        result = await db.execute(stmt)
        return [TestCaseResponse.model_validate(sc) for sc in result.scalars().all()]


@router.post("/duplicates", dependencies=[Depends(require_permissions("tests:read"))])
async def check_duplicates(
    body: DuplicateCheckRequest,
    limit: int = Query(5, ge=1, le=20),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Detect potential duplicate test cases using embedding similarity threshold (>= 0.85)."""
    from app.core.ai.embedding import embedding_service
    
    text_repr = body.title
    if body.description:
        text_repr += f" {body.description}"
        
    embedding = embedding_service.generate_embedding(text_repr)
    
    try:
        distance_expr = TestCase.embedding.cosine_distance(embedding)
        stmt = (
            select(TestCase, (1.0 - distance_expr).label("similarity"))
            .where(
                TestCase.tenant_id == current_user.tenant_id,
                TestCase.embedding.isnot(None),
                TestCase.archived_at.is_(None)
            )
            .order_by(distance_expr)
            .limit(limit)
        )
        result = await db.execute(stmt)
        rows = result.all()
        
        duplicates = []
        for tc, similarity in rows:
            if similarity >= 0.85:
                duplicates.append({
                    "test_case": TestCaseResponse.model_validate(tc),
                    "similarity_score": round(float(similarity), 4)
                })
        return {"duplicates": duplicates}
    except Exception as e:
        # Fallback on DB error
        import logging
        logging.getLogger(__name__).warning("Duplicate detection query failed: %s", e)
        # Simply return exact/fuzzy title match
        stmt = (
            select(TestCase)
            .where(
                TestCase.tenant_id == current_user.tenant_id,
                TestCase.title.ilike(f"%{body.title}%"),
                TestCase.archived_at.is_(None)
            )
            .limit(limit)
        )
        result = await db.execute(stmt)
        cases = result.scalars().all()
        return {
            "duplicates": [
                {"test_case": TestCaseResponse.model_validate(c), "similarity_score": 0.9}
                for c in cases
            ]
        }
