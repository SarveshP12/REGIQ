"""Export service for generating Excel, CSV, and premium PDF documents."""

import io
from fastapi.responses import StreamingResponse
import openpyxl


def generate_excel_export(title: str, columns: list[str], query_results: list) -> StreamingResponse:
    """Generate an in-memory premium Excel streaming response."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]

    # Excel styling colors (sleek slate/blue corporate layout)
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell_font = Font(name="Segoe UI", size=10)
    border_side = Side(style="thin", color="D9D9D9")
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # Write headers
    ws.append(columns)
    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Write data rows
    for row_idx, row in enumerate(query_results, start=2):
        ws.append([getattr(row, col, str(getattr(row, col, ""))) for col in columns])
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = cell_font
            cell.border = cell_border
            if col_num in (3, 4, 5, 7):  # Status, Criticality, Automation, Version
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    filename = f"{title.lower().replace(' ', '_')}.xlsx"
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def generate_pdf_export(title: str, columns: list[str], query_results: list) -> StreamingResponse:
    """Generate a high-fidelity PDF document streaming response using ReportLab."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        # Graceful basic PDF byte stream fallback if ReportLab not loaded
        buf = io.BytesIO()
        buf.write(
            b"%PDF-1.4\n1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
            b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
            b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
            b"/Contents 4 0 R >>\nendobj\n4 0 obj\n<< /Length 55 >>\n"
            b"stream\nBT\n/F1 18 Tf\n50 700 Td\n(REGIQ PDF Export - Offline Fallback) Tj\nET\nendstream\nendobj\n"
            b"xref\n0 5\n0000000000 65535 f\n0000000009 00000 n\n0000000057 00000 n\n"
            b"0000000112 00000 n\n0000000213 00000 n\ntrailer\n<< /Size 5 /Root 1 0 R >>\n"
            b"startxref\n319\n%%EOF"
        )
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={title.lower().replace(' ', '_')}.pdf"},
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "DocTitle",
        parent=styles["Heading1"],
        fontSize=18,
        leading=22,
        textColor=colors.HexColor("#1A365D"),
        spaceAfter=15,
    )

    header_style = ParagraphStyle(
        "TableHeader",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
        fontName="Helvetica-Bold",
        textColor=colors.white,
    )

    cell_style = ParagraphStyle(
        "TableCell",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#2D3748"),
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 8))

    # Format table data
    table_data = []
    table_data.append([Paragraph(col.replace("_", " ").upper(), header_style) for col in columns])

    for row in query_results:
        row_cells = []
        for col in columns:
            val = getattr(row, col, str(getattr(row, col, "")))
            if val is None:
                val = ""
            row_cells.append(Paragraph(str(val), cell_style))
        table_data.append(row_cells)

    # Calculate column widths based on printable document width
    col_width = (doc.width) / len(columns)
    t = Table(table_data, colWidths=[col_width] * len(columns))

    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A365D")),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ("TOPPADDING", (0, 0), (-1, 0), 6),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
                ("TOPPADDING", (0, 1), (-1, -1), 5),
            ]
        )
    )

    elements.append(t)
    doc.build(elements)
    buffer.seek(0)

    filename = f"{title.lower().replace(' ', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
